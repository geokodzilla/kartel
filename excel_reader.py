# -*- coding: utf-8 -*-

from openpyxl import load_workbook, Workbook
from collections import OrderedDict
from app_blad import AppBlad
import os
import string

class ExcelReader(object):

    KOL = {n: i for n, i in enumerate(string.ascii_uppercase, 1)}
    
    def __init__(self, filename, numer_arkusza='1', wiersz_naglowka='1', wiersz_danych='2', kolumna_danych='1', id=['3', '4']):
        self.filename = filename
        self.numer_arkusza = numer_arkusza
        self.wiersz_naglowka = wiersz_naglowka
        self.wiersz_danych = wiersz_danych
        self.kolumna_danych = kolumna_danych
        self.id = id
        self._data_validation()
        
    def _data_validation(self):
        """
        metoda sprawdzająca dane i konwertująca je do postaci numerycznej
        """
        try:
            self.numer_arkusza = int(self.numer_arkusza)
            self.wiersz_naglowka = int(self.wiersz_naglowka)
            self.wiersz_danych = int(self.wiersz_danych)
            self.kolumna_danych = int(self.kolumna_danych)
            self.id = [int(i) if i != '' else i for i in self.id]
        except ValueError:
            raise AppBlad('Błędne parametry wczytania pliku')
                  
    def read(self):
        """
        metoda wczytująca dane z arkusza pliku xlsx, zwraca dane nagłówka
        oraz słownik z danymi wraz zachowaną kolejnością wczytywanych wierszy
        """
        wb = load_workbook(self.filename, data_only=True)
        sn = wb.sheetnames # nazwy arkuszy
        arkusz = wb[sn[self.numer_arkusza-1]] # wybieramy z listy arkusz wg kolejności
        last_col = len(arkusz[str(self.wiersz_naglowka)]) # ustalamy pozycję ostatniej kolumny
        data = OrderedDict()
        try:
            col = self.KOL[self.kolumna_danych]
            last_row = len(arkusz[col])
        except KeyError:
            raise AppBlad('Kolumny poza zakresem')
        else:
            for i, row in enumerate(arkusz.iter_rows(min_row=self.wiersz_danych, max_col=last_col, max_row=last_row), 1): 
                if len(self.id) == 2: # ID dla obrebu i budynku w osobnych kolumnach
                    id_obr = str(row[self.id[0]-1].value)
                    id_bud = str(row[self.id[1]-1].value)
                    id = '-'.join((id_obr, id_bud))
                elif len(self.id) == 1 and self.id != ['']: # id dla budynku w jednej kolumnie
                    id = str(row[self.id[0]-1].value)
                else: # gdy ID nie jest określone id zgodne z numerem wiersza
                    id = str(i)
                data[id] = [str(cell.value) for cell in row]
            header = []
            for row in arkusz.iter_rows(min_row=self.wiersz_naglowka, max_col=last_col, max_row=self.wiersz_danych-1):
                header.append([str(cell.value) for cell in row])
        return data, header
         