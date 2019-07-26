# -*- coding: utf-8 -*-

from openpyxl import load_workbook, Workbook
from collections import OrderedDict
from openpyxl.styles import colors
from openpyxl.styles import Font, Color, PatternFill
from openpyxl.styles.borders import Border, Side
from app_blad import AppBlad

###############################################################################
# STYLES
###############################################################################



class ExcelWriter(object):
    
    # styl ramki
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    # style poszczególnych rodzajów zmian
    styles_dict = {'U': {'font': Font(color='00000000', strike=True),
                         'fill': PatternFill(fill_type='solid', start_color='00FF0000', end_color='00FF0000')
                         },
                   'D': {'font': Font(color='00000000', bold=True, underline='single'),
                         'fill': PatternFill(fill_type='solid', start_color='0000FF00', end_color='0000FF00')},
                   'M': {'fontM': Font(color='000000FF', italic=True, bold=True),
                         'fillM': PatternFill(fill_type='solid', start_color='00808080', end_color='00808080'),
                         'font': Font(color='00000000'),
                         'fill': PatternFill(fill_type='solid', start_color='00FFFFFF', end_color='00FFFFFF')
                         }
                   }



    def __init__(self, header, mod):
        self.header = header
        self.mod = mod


    def write(self, file):
        """
        metoda zapisująca porównanie danych do pliku z raportem
        """
        wb = Workbook()
        ws = wb.active
        ws.title = 'Raport'
        for nr, row in enumerate(self.header, 1):
            for nrc, col in enumerate(row, 1):
                if col == 'None': col = ''
                ws.cell(row=nr, column=nrc, value=col)
        data_row = len(self.header) + 1
        for nr, row in enumerate(self.mod, data_row):
            typ, col_d = row
            for nrc, col in enumerate(col_d, 1):
                if col == 'None': col = ''
                cell = ws.cell(row=nr, column=nrc, value=col)
                if '->' in col:
                    cell.font = self.styles_dict[typ]['fontM']
                    cell.fill = self.styles_dict[typ]['fillM']
                    cell.border = self.thin_border
                else:
                    cell.font = self.styles_dict[typ]['font']
                    cell.fill = self.styles_dict[typ]['fill']
                    cell.border = self.thin_border
        try:
            wb.save(file)
        except PermissionError:
            raise AppBlad('Plik aktualnie w użyciu')